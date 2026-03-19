import openpyxl
from pathlib import Path
from .base import BaseImporter


# Map sheet names to election type slugs
SHEET_TYPE_MAP = {
    'općinsko vijeće': ('local_municipal_council', 'Općinsko vijeće', True),
    'načelnik': ('local_mayor', 'Načelnik', False),
    'gradsko vijeće': ('local_city_council', 'Gradsko vijeće', True),
    'gradonačelnik': ('local_city_mayor', 'Gradonačelnik', False),
    'županijska skupština': ('local_county_assembly', 'Županijska skupština', True),
    'župan': ('local_county_prefect', 'Župan', False),
    'skupština grada zagreba': ('local_city_council', 'Gradsko vijeće', True),
    'gradonačelnik grada zagreba': ('local_city_mayor', 'Gradonačelnik', False),
}

# Partial match for zamjenik sheets
ZAMJENIK_PATTERNS = [
    ('zamjenik načelnika', 'local_deputy_mayor', 'Zamjenik načelnika', False),
    ('zamjenik gradonačelnika', 'local_deputy_city_mayor', 'Zamjenik gradonačelnika', False),
    ('zamjenik župana', 'local_deputy_county_prefect', 'Zamjenik župana', False),
]

GEO_COLS = 13  # columns 0-12


def classify_sheet(sheet_name):
    """Return (slug, display_name, is_list_election) for a sheet name."""
    lower = sheet_name.lower().strip()
    if lower in SHEET_TYPE_MAP:
        return SHEET_TYPE_MAP[lower]
    for prefix, slug, name, is_list in ZAMJENIK_PATTERNS:
        if lower.startswith(prefix):
            return (slug, name, is_list)
    return None


class LocalImporter(BaseImporter):
    """Import local election results from Excel files.

    Each file has 2-4 sheets, each sheet is a different election subtype.
    Row 1: title row (skip), Row 2: header, Row 3+: data.

    For council elections (vijeće/skupština): columns from 13 onward are list names.
    For executive elections (načelnik/gradonačelnik/župan): columns from 13 are candidate names.
    """

    DATA_DIR = Path('/Users/stjepanmudronja/Documents/projekt_izbori/files/Rezultati_lokalni_izbori_2025')

    def run(self):
        for round_num, round_dir in [(1, 'krug-1'), (2, 'krug-2')]:
            round_path = self.DATA_DIR / round_dir
            if not round_path.exists():
                continue
            self.log(f"Importing local elections round {round_num}...")
            files = sorted(round_path.rglob('*.xlsx'))
            self.log(f"  Found {len(files)} files")
            for i, filepath in enumerate(files):
                self._import_file(filepath, round_num)
                if (i + 1) % 50 == 0:
                    self.log(f"  Processed {i + 1}/{len(files)} files")
            self.log(f"  Round {round_num} complete: {len(files)} files")

    def _import_file(self, filepath, round_num):
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        except Exception as e:
            self.log(f"  ERROR opening {filepath.name}: {e}")
            return

        for sheet_name in wb.sheetnames:
            info = classify_sheet(sheet_name)
            if info is None:
                self.log(f"  Unknown sheet type '{sheet_name}' in {filepath.name}, skipping")
                continue

            slug, display_name, is_list_election = info
            ws = wb[sheet_name]
            self._import_sheet(ws, filepath, round_num, slug, display_name, is_list_election)

        wb.close()

    def _import_sheet(self, ws, filepath, round_num, slug, display_name, is_list_election):
        election_type = self.get_or_create_election_type(slug, display_name, parent_slug='local')
        election = self.get_or_create_election(election_type, 2025, f'{display_name} 2025')
        election_round = self.get_or_create_round(election, round_num)

        # Read header (row 2)
        header_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        if not header_row:
            return
        header = list(header_row[0])

        # Get list/candidate names from column 13 onward
        entry_names = []
        for i in range(GEO_COLS, len(header)):
            name = header[i]
            if name is not None:
                name = str(name).strip()
                if name:
                    entry_names.append((i, name))

        if not entry_names:
            return

        # Create electoral lists (and candidacies for executive elections)
        list_data = []
        for col_idx, name in entry_names:
            el_list = self.get_or_create_electoral_list(election_round, name)
            candidacy = None
            if not is_list_election:
                # Executive election: each entry is a candidate
                person = self.get_or_create_person(name)
                candidacy = self.get_or_create_candidacy(person, el_list, 1)
            list_data.append((col_idx, el_list, candidacy))

        # Import data rows (starting from row 3)
        row_count = 0
        for row_data in ws.iter_rows(min_row=3, values_only=True):
            row = list(row_data)
            if len(row) < GEO_COLS or row[0] is None:
                continue

            county_code = str(row[0]).strip()
            county_name = str(row[1]).strip() if row[1] else ''
            muni_type = str(row[2]).strip() if row[2] else ''
            muni_name = str(row[3]).strip() if row[3] else ''
            ps_number = str(row[4]).strip() if row[4] else ''
            ps_name = str(row[5]).strip() if row[5] else ''
            ps_location = str(row[6]).strip() if row[6] else ''
            ps_address = str(row[7]).strip() if row[7] else ''

            registered = self.parse_int(row[8])
            cast = self.parse_int(row[9])
            valid = self.parse_int(row[11])
            invalid = self.parse_int(row[12])

            county = self.get_or_create_county(county_code, county_name)
            municipality = self.get_or_create_municipality(county, muni_name, muni_type)
            polling_station = self.get_or_create_polling_station(
                municipality, ps_number, ps_name, ps_location, ps_address
            )

            self.create_turnout(election_round, polling_station, registered, cast, valid, invalid)

            for col_idx, el_list, candidacy in list_data:
                votes = self.parse_int(row[col_idx]) if col_idx < len(row) else 0
                self.create_list_result(el_list, polling_station, votes)
                if candidacy:
                    self.create_candidate_result(candidacy, polling_station, votes)

            row_count += 1

        self.flush_all()
